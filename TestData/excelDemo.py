import openpyxl

book = openpyxl.load_workbook("C:/Users/DeadShadow/Documents/PythonDemoTestData.xlsx")
sheet = book.active

Dict = {}           #   Empty Dictionary to store Data

#if we want to get cell from that sheet with which row and which column
cell = sheet.cell(row=1, column=2)
print(cell.value)

# sheet['b5'].value = "Jitendra"
# print(sheet['b5'].value)

sheet.cell(row=2, column=3).value = "Kute"
print(sheet.cell(row=2, column=3).value)

print(sheet.max_row)
print(sheet.max_column)

# to fetch complete data from the table and specific data also - both example

for i in range(1, sheet.max_row+1):
    if sheet.cell(row=i, column=1).value == "Testcase2":    #if we want data with specific condition
        for j in range(1,sheet.max_column+1):
            print(sheet.cell(row=i, column=j).value)

# lets save this data into Dictonary
#   1) create Empty Dictionary --> Dict

for i in range(1, sheet.max_row+1):
    if sheet.cell(row=i, column=1).value == "Testcase2":    #if we want data with specific condition
        for j in range(1,sheet.max_column+1):
            Dict[sheet.cell(row=1, column=j).value] = sheet.cell(row=i, column=j).value

print(Dict)

