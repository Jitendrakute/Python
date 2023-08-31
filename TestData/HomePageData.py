import openpyxl

class HomePageData:

    test_HomePage_data = [{"firstname":"jitendra", "email":"Jitendra@123", "Pass":"123"}, {"firstname":"sukh", "email":"Sukh@123", "Pass":"456"}]

    @staticmethod           #avoid creation of class object, in this situation there is no need to use self parameter below
    def getTestData(testcase_name):

        Dict = {}  # Empty Dictionary to store Data
        book = openpyxl.load_workbook("C:/Users/DeadShadow/Documents/PythonDemoTestData.xlsx")
        sheet = book.active

        for i in range(1, sheet.max_row + 1):
            if sheet.cell(row=i, column=1).value == testcase_name:  # if we want data with specific condition but in this condion we pass parameter and request specific data using conftest file
                for j in range(2, sheet.max_column + 1):
                    Dict[sheet.cell(row=1, column=j).value] = sheet.cell(row=i, column=j).value

        #print(Dict)
        return[Dict]        #to binding data like dictionary and send it to form
